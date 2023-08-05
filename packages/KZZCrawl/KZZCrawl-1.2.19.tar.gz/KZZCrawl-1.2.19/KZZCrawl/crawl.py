#coding=utf-8

import os
import math
import random
from KZZCrawl import info
import json
import yaml
import time
import requests
from urllib import parse
from datetime import datetime
from datetime import timedelta
from openpyxl import Workbook
from openpyxl.styles import numbers, Alignment
from openpyxl import load_workbook


rootPath=r'./'

class Excel():
    #初始化
    def __init__(self):
        self.path = ''
    # 创建或打开Excel
    def open(self, path, flag=False):
        self.path = path.encode('utf-8').decode('utf-8')
        if os.path.exists(self.path) and flag:
            self.wb=load_workbook(self.path)
        else:
            self.wb=Workbook()
        self.sheet_init('Sheet')
        # 初始化工作表
    def sheet_init(self, name):
        self.ws = self.wb[name] if name in self.wb.sheetnames else self.wb.create_sheet(name.encode('utf-8').decode('utf-8'))
        # 设置列宽
        # self.ws.column_dimensions.fitToWidth=False
        # return
        self.ws.column_dimensions['A'].width = 15
        self.ws.column_dimensions['B'].width = 15
        self.ws.column_dimensions['C'].width = 15
        self.ws.column_dimensions['D'].width = 15
        self.ws.column_dimensions['E'].width = 15
        self.ws.column_dimensions['F'].width = 15
        self.ws.column_dimensions['G'].width = 15
        self.ws.column_dimensions['H'].width = 15
        self.ws.column_dimensions['I'].width = 15
        self.ws.column_dimensions['J'].width = 15
        self.ws.column_dimensions['K'].width = 15
        self.ws.column_dimensions['L'].width = 15
        self.ws.column_dimensions['M'].width = 15
        self.ws.column_dimensions['N'].width = 15
        self.ws.column_dimensions['O'].width = 15
        self.ws.column_dimensions['P'].width = 15
        self.ws.column_dimensions['Q'].width = 15
        self.ws.column_dimensions['R'].width = 15
        self.ws.column_dimensions['S'].width = 15
        self.ws.column_dimensions['T'].width = 15
        self.ws.column_dimensions['U'].width = 15
        self.ws.column_dimensions['V'].width = 15
        self.ws.column_dimensions['W'].width = 15
        self.ws.column_dimensions['X'].width = 15
        self.ws.column_dimensions['Y'].width = 15
        self.ws.column_dimensions['Z'].width = 15

    # 按行写入Excel
    def write_row(self, row):
        self.ws.append(row)
    # 关闭Excel
    def close(self):
        self.wb.close()
    # 格式化输出
    def format(self):
        for row in self.ws.rows:
            for i in range(self.ws.min_column, self.ws.max_column):
                row[i].alignment = Alignment(horizontal='right')
                # row[i].number_format = numbers.FORMAT_NUMBER_COMMA_SEPARATED1 if i<0 else '#,###'
    # 保存并关闭Excel
    def save(self):
        self.wb.save(filename=self.path)
        # self.wb.close()


class ChoiceStock():
    def __init__(self):
        self.urls = {}                      # 目标网页
        self.headers = {}                   # Get/Post请求头
        self.param = {}
        self.excel = Excel()  # Excel对象
        self.stamp = self.utc2local(self.local2utc(datetime.now())).strftime('%Y%m%d')  # #北京时区时间戳转换为字符串
        self.cookie=requests.cookies.RequestsCookieJar()
        self.session = requests.session()

    # 从配置文件中更新登录链接信息
    def update_info(self):
        self.urls = info.loginUrls                                                  #http地址
        self.headers = info.loginHeaders
        self.param = info.loginParam
        self.cookie.set("cookie", info.loginCookie)
        self.session.cookies.update(self.cookie)

    # 发送Get请求
    def requests_get(self, url, data=None):
        try:
            time.sleep(random.random()*1+0.2)     #0-1区间随机数
            url = url if data is None else url+parse.urlencode(data)
            url=url.replace('%2C',',').replace('%3A',':').replace('%2B','+')
            print(self.urls['proxies']['https'] + ' --> ' + url)
            # response = requests.get(url, timeout=10)
            response = requests.get(url, proxies=self.urls['proxies'], timeout=10)
            # self.session.keep_alive = False
            # response = self.session.get(url, proxies=self.urls['proxies'], timeout=10)
            # response.encoding = 'utf-8'
            return response if response.status_code == requests.codes.ok else None
        except:
            return None
    # 保存htmlPage
    def write_html_page(self, path, page):
        f=open(path, 'w')
        f.write(page.encode('utf-8'))
        f.close()
    # UTC转本地（北京）
    def utc2local(self, utc_st):
        offset = timedelta(hours=8)
        local_st = utc_st + offset
        return local_st
    # 本地转UTC
    def local2utc(self, local_st):
        time_struct = time.mktime(local_st.timetuple())
        utc_st = datetime.utcfromtimestamp(time_struct)
        return utc_st
    # 字符串转换成datetime
    def strptime(self, str):
        return datetime.strptime(str, '%Y-%m-%d')
    # datetime转换为字符串
    def strftime(self, time):
        return time.strftime('%Y%m%d')
    #债券指数
    def crawl1(self):
        url='http://68.push2.eastmoney.com/api/qt/clist/get?pn=1&pz=2000&po=1&np=1&fltt=2&invt=2&fid=f3&fs=i:1.000012,i:1.000013,i:1.000022,i:1.000061,i:0.395021,i:0.395022,i:0.395031,i:0.395032,i:0.399481&fields=f1,f2,f3,f4,f5,f6,f7,f8,f9,f10,f12,f13,f14,f15,f16,f17,f18,f20,f21,f23,f24,f25,f26,f22,f33,f11,f62,f128,f136,f115,f152'
        page=self.requests_get(url)
        mk_json=json.loads(page.text)
        self.excel.write_row(['序号','代码','名称','最新价','涨跌额','涨跌幅','今开','最高','最低','昨收','成交量（手）','成交额'])
        for i in range(0,len(mk_json['data']['diff'])):
            mk=mk_json['data']['diff'][i]
            print(mk['f14'])        #名称
            row=[i+1,mk['f12'],mk['f14'],mk['f2'],mk['f4'],mk['f3'],mk['f17'],mk['f15'],mk['f16'],mk['f18'],mk['f5'],mk['f6']]
            self.excel.write_row(row)

    def crawl2(self):
        self.excel.sheet_init('可转债比价')
        page=self.requests_get(self.urls['url'], self.param['com'])
        mk_json=json.loads(page.text)
        #查询剩余分页并追加到字典
        numNode = mk_json['data']['total']      #记录总数
        numPage = math.ceil(numNode / self.param['com']['pz'])      #页码数
        for i in range(1, numPage):   #跳过第一页
            try:
                self.param['com']['pn'] = i + 1
                page = self.requests_get(self.urls['url'], self.param['com'])
                mk_json['data']['diff'].extend(json.loads(page.text)['data']['diff'])
            except:
                continue
        self.excel.write_row(['序号','转债代码','转债名称','最新价','涨跌幅','正股代码','正股名称','最新价','涨跌幅','转股价','转股价值','转股溢价率','纯债溢价率','回售触发价','强赎触发价','到期赎回价','纯债价值','开始转股日','上市日期','申购日期'])
        for i in range(0,len(mk_json['data']['diff'])):
            mk=mk_json['data']['diff'][i]
            print(mk['f14'])        #名称
            row=[i+1,mk['f12'],mk['f14'],mk['f2'],mk['f3'],mk['f232'],mk['f234'],mk['f229'],mk['f230'],mk['f235'],mk['f236'],mk['f237'],mk['f238'],mk['f239'],mk['f240'],mk['f241'],mk['f227'],mk['f242'],mk['f26'],mk['f243']]
            self.excel.write_row(row)

    def parseFont(self, key, FontMapping):
        for item in FontMapping:
            if item['code']==key:
                return str(item['value'])
        return ''
    def parseFont2(self, font, FontMapping):
        values = ''
        dots=font.split('.')
        #整数部分
        keys=dots[0].split(';')
        for key in keys:
            value=self.parseFont(key+';', FontMapping)
            values+=value
        #小数部分
        if len(dots)>1:
            values += '.'
            keys=dots[1].split(';')
            for key in keys:
                value=self.parseFont(key+';', FontMapping)
                values += value
        return float(values if values != '' else 0)

    def crawl3(self):
        self.excel.sheet_init('可转债数据一览表')
        page=self.requests_get(self.urls['kzz'], self.param['kzz'])
        mk_json=yaml.safe_load(page.text.replace(':',': '))
        #查询剩余分页并追加到字典
        numPage=mk_json['pages']
        for i in range(1, numPage):   #跳过第一页
            try:
                self.param['kzz']['p'] = i + 1
                page = self.requests_get(self.urls['kzz'], self.param['kzz'])
                mk_json['data'].extend(yaml.safe_load(page.text.replace(':',': '))['data'])
            except:
                continue
        self.excel.write_row(['债券代码','债券名称','申购日期','申购代码','申购上限（万元）','正股代码','正股简称','正股价','转股价','转股价值','债现价','转股溢价率','股权登记日','每股配售额','发行规模（亿元）','中签号发布日','中签率','上市时间','PB','HSCFJ','QSCFJ','MEMO'])
        for i in range(0,len(mk_json['data'])):
            mk=mk_json['data'][i]
            print(mk['SNAME'])        #名称
            try:
                if '-' in [mk['ZGJ'], mk['SWAPPRICE']]:
                    zgjz = '-'
                    zgyjl = '-'
                else:
                    zgjz = float(mk['ZGJ']) / float(mk['SWAPPRICE']) * 100
                    zgyjl = float(mk['SWAPPRICE']) / float(mk['ZGJ']) - 1
            except:
                print(i)
                zgjz = '-'
                zgyjl = '-'
            #映射加密的数字
            mgpse=self.parseFont2(mk['FSTPLACVALPERSTK'],mk_json['font']['FontMapping'])
            fxgm=self.parseFont2(mk['AISSUEVOL'],mk_json['font']['FontMapping'])
            #写入Excel
            row=[mk['BONDCODE'],mk['SNAME'],mk['STARTDATE'],mk['CORRESCODE'],mk['PARVALUE'],mk['SWAPSCODE'],mk['SECURITYSHORTNAME'],mk['ZGJ'], mk['SWAPPRICE'],zgjz,mk['ISSUEPRICE'],zgyjl,mk['GDYX_STARTDATE'],mgpse,fxgm,mk['ZQHDATE'],mk['LUCKRATE'],mk['LISTDATE'],mk['PB'],mk['HSCFJ'],mk['QSCFJ'],mk['MEMO']]  #,mk['ZGJZGJ'],mk['ZGJZGJJZ'],mk['ZQNEW'],mk['YJL']
            self.excel.write_row(row)

    # 爬取数据并保存
    def Imatate(self):
        # 同步配置参数
        self.update_info()
        # 打开Excel
        self.excel.open(rootPath + 'KZZ' + self.stamp + '.xlsx')
        self.crawl1()
        self.crawl2()
        self.crawl3()
        # 关闭并保存Excel
        self.excel.format()
        self.excel.save()

# 主程序入口
if __name__ == '__main__':
    spider = ChoiceStock()
    spider.Imatate()

    print('TS:'+spider.utc2local(spider.local2utc(datetime.now())).strftime('%Y-%m-%d %H:%M:%S'))


