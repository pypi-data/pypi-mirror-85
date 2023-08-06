#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""

    thegmu_im_catalog_excel.py
    ~~~~~~~~~~~~~~~~~~~~~~~~~~

    Library that loads and saves the catalog to an Excel spreadsheet format.
    New columns with image links to a thumbnail and original image
    are added as well as new column for individual file commands.

    #. file:///thumbnail/image.jpg column added.
    #. file:///image.jpg column added.
    #. command column added.

"""

import datetime
import os
from collections import OrderedDict
from openpyxl import drawing as OpenypxlDrawing
from openpyxl import utils as OpenpyxlUtils
from openpyxl import Workbook as OpenpyxlWorkbook
from openpyxl.styles import Alignment as OpenpyxlAlignment
from PIL import Image


class TheGMUImageManageCatalogExcel():
    """Load and save the catalog to an Excel format.
    Includes two new image link columns and a command column.
    """

    CATALOG_EXCEL_EXT = 'xlsx'
    CATALOG_EXCEL_COLS = ('thumbnail',
                          'original',
                          'note',
                          'delete',
                          'move',
                          'rename',
                          'tags', )

    CATALOG_EXCEL_USER_COLS = ('note',
                               'delete',
                               'move',
                               'rename',
                               'tags', )

    CATALOG_EXCEL_SHEET_KEY = 'Catalog'
    CATALOG_EXCEL_THUMBNAIL_DIR = "thumbnails"
    CATALOG_EXCEL_THUMBNAIL_HEIGHT = 100
    CATALOG_EXCEL_THUMBNAIL_WIDTH = 300
    CATALOG_EXCEL_MAX_ROWS = 1000

    CATALOG_EXCEL_CELL_HEIGHT = 80
    CATALOG_EXCEL_CELL_WIDTH = 20

    # Values like 3.5 are in inches.
    # Values were set by opening sample spreadsheets and using the
    # size in inches reported by Excel.
    CATALOG_CELL_SIZES = OrderedDict([('thumbnail', 3.5),
                                      ('original', None),
                                      ('note', 2),
                                      ('delete', 0.7),
                                      ('move', None),
                                      ('rename', None),
                                      ('tags', None),
                                      ('file_name', None),
                                      ('format', 0.65),
                                      ('WxH', 1.3),
                                      ('size', 0.9),
                                      ('date', 2),
                                      ('epoch', 1.2),
                                      ('md5sum', 3.6),
                                      ('ext', 0.5)])

    def _catalog_emit_excel_init(self):
        """Initialize the thumbnail directory if it doesn't exist."""

        thumbnail_path = self.CATALOG_EXCEL_THUMBNAIL_DIR

        if (not os.path.exists(thumbnail_path)):
            os.mkdir(thumbnail_path)

    @staticmethod
    def _catalog_emit_excel_original(_sheet, cell, work_file):
        """original creates a hyperlink to the original file."""

        cell.value = None
        cell.hyperlink = work_file

    def _catalog_emit_excel_thumbnail(self, sheet, cell, work_file):
        """Create a thumbnail in the thumbnail directory and
        set the cell image."""

        thumbnail_path = self.CATALOG_EXCEL_THUMBNAIL_DIR
        thumbnail_height = self.CATALOG_EXCEL_THUMBNAIL_HEIGHT
        thumbnail_width = self.CATALOG_EXCEL_THUMBNAIL_WIDTH

        image_thumbnail_path = os.path.join(thumbnail_path, work_file)

        if (not os.path.exists(image_thumbnail_path)):

            with Image.open(work_file) as work_img:

                (width, height) = work_img.size
                if ((height <= thumbnail_height) and
                        (width <= thumbnail_width)):
                    image_thumbnail_path = work_file
                else:
                    new_width = int((thumbnail_height / height) * width)
                    if (new_width > thumbnail_width):
                        new_height = int((thumbnail_width / width) *
                                         height)
                        new_width = thumbnail_width
                    else:
                        new_height = thumbnail_height

                    work_img.thumbnail((new_width, new_height))
                    work_img.save(image_thumbnail_path)

        sheet_img = OpenypxlDrawing.image.Image(image_thumbnail_path)
        sheet_img.anchor = cell.coordinate
        sheet.add_image(sheet_img)
        self.log.debu("thumbnail: %s" % (image_thumbnail_path, ))

    def _catalog_emit_excel(self, start_row_count=0):
        """Write the catalog Excel file from previously seeded stat and
        imagemagick data."""

        catalog_file = self.data['catalog']['catalog_file']
        working_directory = self.data['catalog']['working_directory']

        self.log.debu('cwd: %s' % (os.getcwd()))
        catalog_excel_file = self.replace_file_ext(catalog_file, '')

        book_count = int(start_row_count / self.CATALOG_EXCEL_MAX_ROWS) + 1
        new_ext = "%s.%02d.%s" % (datetime.date.today(), book_count,
                                  self.CATALOG_EXCEL_EXT)

        catalog_excel_file += new_ext

        if (os.path.exists(catalog_excel_file)):
            self.log.warn("'" + catalog_excel_file + "': " +
                          "cowardly refusing to overwrite existing Excel file"
                          " that you may have updated.")

        else:

            os.chdir(working_directory)
            book = OpenpyxlWorkbook()
            sheet = book.active
            sheet.title = self.CATALOG_EXCEL_SHEET_KEY

            self._catalog_emit_excel_init()
            self._catalog_excel_set_rows()

            row_count = self._catalog_excel_set_sheet(sheet, start_row_count)

            try:
                book.save(catalog_excel_file)
            except PermissionError:
                self.log.warn("%s Excel file permission denied" %
                              (catalog_excel_file))
            else:
                self.log.prog("%s file created." % (catalog_excel_file, ))

            start_row_count += row_count

            self.cd_start_dir()

            if (start_row_count < len(self.data['catalog']['excel'])):
                self._catalog_emit_excel(start_row_count)

    @classmethod
    def _catalog_excel_get_column_width(cls, column_index):
        """Given a column name then return the column width.
        The width is in tenths of inches.
        """

        column_size = list(cls.CATALOG_CELL_SIZES.values())[column_index]

        if (column_size is None):
            return cls.CATALOG_EXCEL_CELL_WIDTH

        column_size *= 10  # Excel sizes are in tenths of an innch

        return column_size

    def _catalog_excel_set_rows(self):
        """Expand the 'current' catalog to include the 'excel'
        columns.
        """
        if ('excel' in self.data['catalog']):
            return

        current = self.data['catalog']['current']

        excel_rows = []

        excel_columns = [""] * len(self.CATALOG_EXCEL_COLS)

        for work_file in current:
            row = excel_columns + list(current[work_file].values())
            excel_rows.append(row)

        self.data['catalog']['excel'] = excel_rows

    def _catalog_excel_set_sheet(self, sheet, start_row_count):
        """Helper function"""

        work_file_index = len(self.CATALOG_EXCEL_COLS)

        self._catalog_excel_set_sheet_header(sheet)
        row_count = 0

        # 'row_index + 2' accounts for 1 header row and spreadheet
        # index starts at 1, not 0
        for row_index, row in enumerate(
                self.data['catalog']['excel'][start_row_count:]):
            sheet.row_dimensions[row_index + 2].height = \
                self.CATALOG_EXCEL_CELL_HEIGHT
            for column_index, col_value in enumerate(row):
                column_width = self._catalog_excel_get_column_width(
                    column_index)

                column_letter = \
                    OpenpyxlUtils.get_column_letter((column_index + 1))
                sheet.column_dimensions[column_letter].width = column_width

                cell_addr = '%s%s' % (column_letter, (row_index + 2))
                sheet[cell_addr].value = col_value
                sheet[cell_addr].alignment = OpenpyxlAlignment(
                    horizontal="left", vertical="top", wrap_text=True)

                if ((column_index < len(self.CATALOG_EXCEL_COLS)) and
                        (self.CATALOG_EXCEL_COLS[column_index] not in
                         self.CATALOG_EXCEL_USER_COLS)):
                    try:
                        method_name = ("_catalog_emit_excel_%s" %
                                       self.CATALOG_EXCEL_COLS[column_index])
                        getattr(self, method_name)(sheet, sheet[cell_addr],
                                                   row[work_file_index])
                    except AttributeError as attr_error:
                        self.log.debu("%s %s" % (method_name, attr_error))
                        msg = ("%s is missing, programming error." %
                               method_name)
                        raise AttributeError(msg) from attr_error

            row_count += 1
            if (row_count >= self.CATALOG_EXCEL_MAX_ROWS):
                break

        return row_count

    def _catalog_excel_set_sheet_header(self, sheet):
        """Set the header outside of row count so the number of rows
        equals MAX_ROWS.
        """
        header = [*self.CATALOG_EXCEL_COLS] + list(self.CATALOG_RECORD.keys())
        get_letter = OpenpyxlUtils.get_column_letter

        for column_index, col_value in enumerate(header):
            sheet["%s1" % get_letter((column_index + 1))].value = col_value
