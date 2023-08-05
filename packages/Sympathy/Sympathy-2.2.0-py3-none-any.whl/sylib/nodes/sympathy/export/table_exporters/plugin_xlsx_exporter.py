# This file is part of Sympathy for Data.
# Copyright (c) 2017, Combine Control Systems AB
#
# Sympathy for Data is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, version 3 of the License.
#
# Sympathy for Data is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with Sympathy for Data.  If not, see <http://www.gnu.org/licenses/>.
import re
import os
import tempfile
import numpy as np
import unicodedata
import openpyxl
import openpyxl.styles
import openpyxl.cell
import datetime
from sylib.export import table as exporttable
from sympathy.api import exceptions
from sympathy.api import qt2 as qt_compat
from sympathy.api.nodeconfig import settings

import sympathy.api


QtGui = qt_compat.import_module('QtGui')
QtWidgets = qt_compat.import_module('QtWidgets')


def _col2ai(i):
    """
    Return corresponding A, B, .. AB notation for column number i.
    """
    chars = []
    c = i % 26
    chars.append(c)
    i = (i - c) // 26

    while i > 0:
        c = (i - 1) % 26
        chars.append(c)
        i = (i - c) // 26

    return ''.join([chr(ord('A') + c_) for c_ in reversed(chars)])


def _range_name(column_name):
    """
    Convert column name to valid name for use in workbook named range.
    crng_ + converted compatible name.
    """
    if isinstance(column_name, bytes):
        column_name = column_name.decode('ascii', errors='ignore')
    return u'crng_{}'.format(
        re.sub('[^a-zA-Z0-9]', '_', unicodedata.normalize('NFD', column_name)))


def create_filled_sheet(workbook, sheet_name, input_table, header):
    column_names = input_table.column_names()
    columns = {key: i for i, key in
               enumerate(column_names)}
    worksheet = workbook.create_sheet(sheet_name)

    # Write data columns.
    if header:
        bold = openpyxl.styles.Font(bold=True)
        columns = []
        for column_name in column_names:
            cell = openpyxl.cell.WriteOnlyCell(
                worksheet, value=column_name)
            cell.font = bold
            columns.append(cell)

        worksheet.append(columns)
        start_row = 1
    else:
        start_row = 0

    end_row = input_table.number_of_rows() + start_row
    warn_about_datetimes = False
    warn_about_timedeltas = False

    output_table = sympathy.api.table.File()

    def mask_unwritable(arr, data, unwritable):
        if np.any(unwritable):
            mask = np.ma.getmaskarray(arr)
            if np.any((~mask) * unwritable):
                return True, np.ma.masked_array(
                    data, mask | unwritable)
        return False, None

    for i, column_name in enumerate(input_table.column_names()):
        kind = input_table.column_type(column_name).kind
        arr = input_table[column_name]
        data = np.ma.getdata(arr)
        new_col = None

        workbook.create_named_range(
            name=_range_name(column_name),
            worksheet=worksheet,
            value="${col}${srow}:${col}${erow}".format(
                sheet=sheet_name,
                col=_col2ai(i),
                srow=start_row + 1,
                erow=end_row),
            scope=workbook.index(worksheet))

        if kind == 'M':

            # All dates before 1900-03-01 are ambiguous because of
            # a bug in Excel which incorrectly treats the year 1900
            # as a leap year. So don't write any such dates to
            # file.

            unwritable = data < datetime.datetime(1900, 3, 1)
            warn, new_col = mask_unwritable(
                arr,
                data,
                unwritable)
            warn_about_datetimes |= warn

        elif kind == 'm':

            # Timedeltas can't be written to excel in general, but
            # if the delta is less than a day, it can be
            # represented as a time.

            warn, new_col = mask_unwritable(
                arr,
                data,
                data >= datetime.timedelta(days=1))
            warn_about_timedeltas |= warn

        if new_col is None:
            output_table.update_column(column_name, input_table)
        else:
            output_table[column_name] = new_col

    for row in output_table.to_rows():
        worksheet.append(row)

    # Give a warning about problematic datetimes and/or timedeltas
    if warn_about_datetimes:
        exceptions.sywarn(
            "Not writing any datetimes before 1900-03-01. All dates "
            "before 1900-03-01 are ambiguous because of a bug in Excel.")
    if warn_about_timedeltas:
        exceptions.sywarn(
            "Not writing any timedeltas bigger than or equal to one full "
            "day. Such timedeltas can't be expressed in Excel.")


class DataExportXLSXWidget(QtWidgets.QWidget):
    filename_changed = qt_compat.Signal()

    def __init__(self, parameter_root, input_list):
        super(DataExportXLSXWidget, self).__init__()
        self._parameter_root = parameter_root
        self._input_list = input_list
        self._init_gui()

    def _init_gui(self):
        vlayout = QtWidgets.QVBoxLayout()

        self._to_sheets = self._parameter_root['to_sheets']
        self._to_sheets_check_box = self._to_sheets.gui()
        self._to_sheets_check_box.valueChanged.connect(
            self._filename_changed)

        self._update_sheets = self._parameter_root['update_sheets']
        self._update_sheets_check_box = self._update_sheets.gui()
        self._update_sheets_check_box.valueChanged.connect(
            self._filename_changed)
        self._update_sheets_check_box.setEnabled(self._to_sheets.value)

        self._table_names = self._parameter_root['table_names']
        self.table_names_gui = self._parameter_root['table_names'].gui()
        self.table_names_gui.valueChanged.connect(self._filename_changed)
        self.table_names_gui.setEnabled(not self._to_sheets.value)

        vlayout.addWidget(self.table_names_gui)
        vlayout.addWidget(self._to_sheets_check_box)
        vlayout.addWidget(self._update_sheets_check_box)
        vlayout.addWidget(self._parameter_root['header'].gui())

        self._to_sheets_check_box.stateChanged[int].connect(
            self._to_sheets_state_changed)

        self.setLayout(vlayout)

    @qt_compat.Slot(int)
    def _to_sheets_state_changed(self, value):
        if self._to_sheets.value:
            self._table_names.value = False
            self.table_names_gui.setEnabled(False)
            self._update_sheets_check_box.setEnabled(True)
        else:
            self.table_names_gui.setEnabled(True)
            self._update_sheets_check_box.setEnabled(False)

    def _filename_changed(self):
        self.filename_changed.emit()


class DataExportXLSX(exporttable.TableDataExporterBase):
    """Exporter for XLSX files."""

    EXPORTER_NAME = 'XLSX'
    FILENAME_EXTENSION = 'xlsx'
    DEFAULT_XLSX_NAME = 'xlsx_filename'

    def __init__(self, parameters):
        super(DataExportXLSX, self).__init__(parameters)

        if 'header' not in parameters:
            parameters.set_boolean(
                'header', value=True, label='Export header',
                description='Export column names')

        if 'to_sheets' not in parameters:
            parameters.set_boolean(
                'to_sheets', label='Export data to sheets',
                description=(
                    'Select to export incoming tables to one file with '
                    'multiple sheets instead of multiple files with one '
                    'sheet.'))

        if 'update_sheets' not in parameters:
            parameters.set_boolean(
                'update_sheets', label='Replace existing sheets',
                description=(
                    'Export data to sheets in an existing file, creating the '
                    'file if it does not exist. Existing sheets are replaced '
                    'by input tables when table name == sheet name, and are '
                    'otherwise preserved.\n\n'
                    'Note that the existing file is re-written and preserved '
                    'sheets are first read and then written, a handling which '
                    'may result in some properties being lost.'))

        if 'table_names' not in parameters:
            parameters.set_boolean(
                'table_names',
                label='Use table names as filenames',
                description='Use table names as filenames')

    def parameter_view(self, input_list):
        return DataExportXLSXWidget(
            self._parameters, input_list)

    def export_data(self, in_data, filename, progress=None):
        """Export Table to XLSX."""
        header = self._parameters['header'].value
        to_sheets = self._parameters['to_sheets'].value
        update_sheets = self._parameters['update_sheets'].value
        update_sheets &= to_sheets
        existing_workbook = False

        table_names = self._parameters['table_names'].value
        if not to_sheets:
            in_data = [in_data]

        if update_sheets:
            try:
                workbook = openpyxl.load_workbook(filename)
                existing_workbook = True
            except Exception:
                workbook = openpyxl.Workbook()
                # Remove default worksheet which would otherwise
                # be created.
                workbook.remove(workbook.active)
        else:
            workbook = openpyxl.Workbook(write_only=True)

        local_worksheet_names = []
        if existing_workbook:
            # Remove all existing local names to work around index issues in
            # openpyxl.

            for worksheet in workbook.worksheets:
                index = workbook.index(worksheet)
                local_worksheet_names.append((
                    worksheet,
                    [workbook.defined_names.get(n, scope=index)
                     for n in workbook.defined_names.localnames(
                        scope=index)]))

            for worksheet, local_names in local_worksheet_names:
                index = workbook.index(worksheet)
                for local_name in local_names:
                    workbook.defined_names.delete(local_name.name, scope=index)

        sheet_names = []

        for i, in_table in enumerate(in_data):
            if table_names:
                sheet_name = os.path.basename(in_table.get_name())
            else:
                sheet_name = in_table.get_name()

            sheet_name = sheet_name or 'Table_{0}'.format(i)

            if len(sheet_name) > 23:
                sheet_name = sheet_name[:23]

            j = 1
            sheet_base = sheet_name
            while sheet_name in sheet_names:
                sheet_name = f'{sheet_base}{j}'
                j += 1

            sheet_names.append(sheet_name)

            if existing_workbook and sheet_name in workbook:
                # Should be removed before starting to create new one
                # because of index issues.
                del workbook[sheet_name]

        for i, (sheet_name, in_table) in enumerate(zip(sheet_names, in_data)):
            create_filled_sheet(
                workbook, sheet_name, in_table, header)

        if existing_workbook:
            # Restore local names that are scoped to existing worksheets.

            worksheets = set(workbook.worksheets)

            for worksheet, local_names in local_worksheet_names:
                if worksheet in worksheets:
                    index = workbook.index(worksheet)
                    for local_name in local_names:
                        local_name.localSheetId = index
                        workbook.defined_names.append(local_name)

        if existing_workbook:
            with tempfile.NamedTemporaryFile(
                    prefix='plugin_xlsx_exporter_', suffix='.xlsx',
                    delete=False,
                    dir=settings()['session_folder']) as f:
                workbook.save(f.name)
            workbook.close()
            os.replace(f.name, filename)
        else:
            workbook.save(filename)
            workbook.close()

    def create_filenames(self, input_list, filename):
        to_sheets = self._parameters['to_sheets'].value
        filename_generator = super(DataExportXLSX, self).create_filenames(
            input_list, filename)
        if to_sheets:
            return [next(iter(filename_generator))]
        return filename_generator

    def cardinality(self):
        to_sheets = self._parameters['to_sheets'].value
        if to_sheets:
            return self.many_to_one
        return self.one_to_one
