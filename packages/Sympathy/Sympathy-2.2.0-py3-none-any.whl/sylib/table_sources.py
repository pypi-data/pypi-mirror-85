# This file is part of Sympathy for Data.
# Copyright (c) 2013, 2016-2017, Combine Control Systems AB
#
# Sympathy for Data is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, version 3 of the License.
#
# Sympathy for Data is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with Sympathy for Data.  If not, see <http://www.gnu.org/licenses/>.
import codecs
import copy
import datetime
import os
import re
import contextlib
from distutils.version import StrictVersion

import numpy as np
import pandas as pd
import openpyxl

from sympathy.api import qt2 as qt_compat
from sympathy.api.exceptions import sywarn, SyDataError
from sympathy.api import table
from sylib.matlab import matlab
from .xl_utils import is_xls, get_xlsx_sheetnames

QtCore = qt_compat.QtCore  # noqa


PANDAS_CSV_READER = pd.io.parsers.read_csv

CSV_FILE_DELIMITERS = {'Tab': '\t',
                       'Comma': ',',
                       'Semicolon': ';',
                       'Space': ' ',
                       'Other': ''}

CODEC_LANGS = {'Western (ASCII)': 'ascii',
               'Western (ISO 8859-1)': 'iso8859_1',
               'Western (ISO 8859-15)': 'iso8859_15',
               'Western (Windows 1252)': 'windows-1252',
               'UTF-16 BE': 'utf_16_be',
               'UTF-16 LE': 'utf_16_le',
               'UTF-16': 'utf_16',
               'UTF-8': 'utf_8'}

CHUNK_ROW_LIMIT = 1000
CHUNK_BYTE_LIMIT = 1 * 1024 * 1024
SNIFF_LIMIT = 5 * 1024 * 1024
ITER_LIMIT = 50
FIND_NR_KEY = r'\d'
FIND_LINE_NR = r'line \d+'
TOO_MANY_COLUMNS = r'[\w\W]*Expected \d+ fields in line (\d+), saw \d+'
END_OF_LINE = r'[\w\W]*EOF inside string starting at (?:row|line) (\d+)'


@contextlib.contextmanager
def open_workbook(filename):
    """Open workbook, on demand - if possible."""
    with contextlib.closing(openpyxl.load_workbook(filename, data_only=True)) as wb:
        yield wb


class PreviewWorker(QtCore.QObject):
    preview_ready = qt_compat.Signal(table.File)

    def __init__(self, import_function):
        super().__init__()
        self._data_table = None
        self._import_function = import_function

    def create_preview_table(self, *args):
        self._data_table = table.File()
        self._import_function(self._data_table, *args)
        self.preview_ready.emit(self._data_table)


class TableSourceModel(QtCore.QObject):
    """Model class, layer between importers and GUIs."""

    update_table = qt_compat.Signal()

    def __init__(self, parameters, fq_infilename, mode):
        super().__init__()
        self._parameters = parameters

        self._fq_infilename = fq_infilename
        self.data_table = None

        self._init_model_common_parameters()

    def _init_model_common_parameters(self):
        """Init common parameters xlsx, csv and mat importers."""
        self.preview_start_row = self._parameters['preview_start_row']
        self.no_preview_rows = self._parameters['no_preview_rows']
        self.data_offset = self._parameters['data_start_row']
        self.data_rows = self._parameters['data_end_row']
        self.data_read_selection = self._parameters['read_selection']

    def _init_model_specific_parameters(self):
        """Method to include the init of special parameters for a
        specific importer.
        """
        pass

    @qt_compat.Slot()
    def collect_preview_values(self):
        """Method that asks the importer for preview data."""
        pass


class TableSourceModelXLS(TableSourceModel):
    """Model layer between GUI and xlsx importer."""

    get_preview = qt_compat.Signal(
        str, bool, int, int, bool, int, int, int)

    def __init__(self, parameters, fq_infilename, mode, valid, multi):
        super().__init__(
            parameters, fq_infilename, mode)
        self.data_table = None
        self._multi = multi
        self._valid = valid
        self._init_model_specific_parameters()

        self._init_preview_worker()

    def _init_model_specific_parameters(self):
        """Init special parameters for xlsx importer."""
        self.headers = self._parameters['headers']
        self.header_row = self._parameters['header_row']
        self.units = self._parameters['units']
        self.unit_row = self._parameters['unit_row']
        self.descriptions = self._parameters['descriptions']
        self.description_row = self._parameters['description_row']
        self.transposed = self._parameters['transposed']
        self.worksheet_name = self._parameters['worksheet_name']
        self.import_all = self._parameters['import_all']

    def _init_preview_worker(self):
        """Collect preview data from xlsx file."""
        if self._valid:
            sheet_names = get_xlsx_sheetnames(self._fq_infilename)
            self.worksheet_name.list = sheet_names
        self._importer = ImporterXLS(self._fq_infilename, False)

        self._preview_thread = QtCore.QThread()
        self._preview_worker = PreviewWorker(self._importer.import_xls)
        self._preview_worker.moveToThread(self._preview_thread)
        self._preview_thread.finished.connect(self._preview_worker.deleteLater)
        self.get_preview.connect(self._preview_worker.create_preview_table)
        self._preview_worker.preview_ready.connect(self.set_preview_table)
        self._preview_thread.start()

        self.collect_preview_values()

    @qt_compat.Slot()
    def collect_preview_values(self):
        """Collect preview data from xlsx file."""
        data_start_row = self.data_offset.value - 1
        data_end_row = data_start_row + self.no_preview_rows.value

        sheet_name = self.worksheet_name.selected
        import_all = self.import_all.value

        transposed = self.transposed.value

        if self.headers.value:
            headers_row = self.header_row.value - 1
        else:
            headers_row = -1
        if self.units.value:
            units_row = self.unit_row.value - 1
        else:
            units_row = -1
        if self.descriptions.value:
            descriptions_row = self.description_row.value - 1
        else:
            descriptions_row = -1
        if self._valid:
            self.get_preview.emit(
                sheet_name, import_all, data_start_row, data_end_row,
                transposed, headers_row, units_row, descriptions_row)
        else:
            self.data_table = table.File()
            self.update_table.emit()

    @qt_compat.Slot(table.File)
    def set_preview_table(self, data_table):
        self.data_table = data_table
        self.update_table.emit()

    def cleanup(self):
        self._preview_thread.quit()
        self._preview_thread.wait()


class TableSourceModelCSV(TableSourceModel):
    """Model layer between GUI and csv importer."""

    get_preview = qt_compat.Signal(int, int, int, int, int, int)

    def __init__(self, parameters, fq_infilename, mode, valid):
        super().__init__(
            parameters, fq_infilename, mode)
        self.data_table = None
        self._csv_source = TableSourceCSV(fq_infilename, valid=valid)
        self._importer = ImporterCSV(self._csv_source)
        self._importer.set_partial_read(True)
        self._csv_delimiters = copy.copy(CSV_FILE_DELIMITERS)
        self._valid = valid
        self._init_model_specific_parameters()

        if self.headers.value is None:
            self.headers.value = self._csv_source.has_header
            if self.headers.value:
                self.data_offset.value += 1

        self._init_preview_worker()

    def _init_model_specific_parameters(self):
        """Init special parameters for csv importer."""
        self.headers = self._parameters['headers']
        self.header_row = self._parameters['header_row']
        self.units = self._parameters['units']
        self.unit_row = self._parameters['unit_row']
        self.descriptions = self._parameters['descriptions']
        self.description_row = self._parameters['description_row']
        self.source_coding = self._parameters['source_coding']
        self.delimiter = self._parameters['delimiter']
        self.custom_delimiter = self._parameters['other_delimiter']
        self.double_quotations = self._parameters['double_quotations']

        self.exceptions_mode = self._parameters['exceptions']

        if self.delimiter.value is None:
            self.delimiter.value = self._csv_source.delimiter
        else:
            self._csv_source.delimiter = self.delimiter.value

        if self._valid:
            if self.source_coding.value is None:
                self.source_coding.value = self._csv_source.encoding
            else:
                self._csv_source.encoding = self.source_coding.value

            if self.custom_delimiter.value is None:
                self.custom_delimiter.value = self._csv_delimiters['Other']
            else:
                self._csv_delimiters['Other'] = self.custom_delimiter.value

            self.encoding_key = [
                key for key, value in CODEC_LANGS.items()
                if value == self.source_coding.value][0]
        else:
            self.encoding_key = [
                key for key, value in CODEC_LANGS.items()
                if value == self._csv_source.DEFAULT_ENCODING][0]

        self.delimiter_key = [
            key for key, value in self._csv_delimiters.items()
            if value == self.delimiter.value][0]

    def _init_preview_worker(self):
        """Collect preview data from xls file."""
        def import_function(*args):
            try:
                self._importer.import_csv(*args, require_num=False)
            except Exception:
                self._importer.import_csv(
                    *args, require_num=False, read_csv_full_rows=True)

        self._preview_thread = QtCore.QThread()
        self._preview_worker = PreviewWorker(import_function)
        self._preview_worker.moveToThread(self._preview_thread)
        self._preview_thread.finished.connect(self._preview_worker.deleteLater)
        self.get_preview.connect(self._preview_worker.create_preview_table)
        self._preview_worker.preview_ready.connect(self.set_preview_table)
        self._preview_thread.start()

        self.collect_preview_values()

    @property
    def delimiter_key(self):
        return self._delimiter_key

    @delimiter_key.setter
    def delimiter_key(self, delimiter_key):
        self._delimiter_key = delimiter_key

    @property
    def encoding_key(self):
        return self._encoding_key

    @encoding_key.setter
    def encoding_key(self, encoding_key):
        self._encoding_key = encoding_key

    @qt_compat.Slot(str)
    def set_new_custom_delimiter(self, value):
        """Set new custom delimiter from GUI."""
        self.custom_delimiter.value = str(value)
        self._csv_delimiters['Other'] = str(value)
        self.set_delimiter('Other')

    @qt_compat.Slot(int)
    def set_double_quotations(self, value):
        self._set_source_double_quotations()
        self.collect_preview_values()

    def _set_source_double_quotations(self):
        self._csv_source.double_quotations = (
            self.double_quotations.value)

    @qt_compat.Slot(str)
    def set_delimiter(self, delimiter_key):
        """Set changed delimiter from GUI."""
        self.delimiter.value = self._csv_delimiters[str(delimiter_key)]
        self._csv_source.delimiter = self.delimiter.value
        self.collect_preview_values()

    @qt_compat.Slot(str)
    def set_encoding(self, encoding_key):
        """Set changed encodings from GUI."""
        self.source_coding.value = CODEC_LANGS[str(encoding_key)]
        self._csv_source.encoding = self.source_coding.value
        self.collect_preview_values()

    @qt_compat.Slot()
    def collect_preview_values(self):
        """Collect preview data from csv file."""
        no_rows = self.no_preview_rows.value
        row_offset = self.data_offset.value - 1

        if self.headers.value:
            headers_row = self.header_row.value - 1
        else:
            headers_row = -1
        if self.units.value:
            units_row = self.unit_row.value - 1
        else:
            units_row = -1
        if self.descriptions.value:
            descriptions_row = self.description_row.value - 1
        else:
            descriptions_row = -1

        if self._csv_source.valid_encoding:
            self.get_preview.emit(
                no_rows, 0, row_offset, headers_row, units_row,
                descriptions_row)
        elif not self._valid:
            self.data_table = table.File()
        else:
            error_string = 'Invalid encoding'
            self.data_table = table.File()
            self.data_table.set_column_from_array(
                'X0', np.array([error_string]))

        self.update_table.emit()

    @qt_compat.Slot(table.File)
    def set_preview_table(self, data_table):
        self.data_table = data_table
        self.update_table.emit()

    def cleanup(self):
        self._preview_thread.quit()
        self._preview_thread.wait()


class TableSourceModelMAT(TableSourceModel):
    """Model layer between GUI and mat importer."""

    get_preview = qt_compat.Signal(int, int, int, int, int, int)

    def __init__(self, parameters, fq_infilename, mode, valid):
        super().__init__(
            parameters, fq_infilename, mode)
        self.data_table = None
        self._valid = valid
        self._mat_source = TableSourceMAT(fq_infilename)
        self._importer = ImporterMAT(self._mat_source)
        self._importer.set_partial_read(True)

        self._init_model_specific_parameters()
        self._init_preview_worker()

    def _init_model_specific_parameters(self):
        """Init special parameters for mat importer."""
        self.preview_start_row = self._parameters['preview_start_row']
        self.no_preview_rows = self._parameters['no_preview_rows']
        self.data_offset = self._parameters['data_start_row']
        self.data_rows = self._parameters['data_end_row']
        self.data_read_selection = self._parameters['read_selection']

    def _init_preview_worker(self):
        """Collect preview data from mat file."""
        def import_function(*args):
            try:
                self._importer.import_mat(*args, require_num=False)
            except Exception:
                self._importer.import_mat(*args, require_num=False,
                                          read_mat_full_rows=True)

        self._preview_thread = QtCore.QThread()
        self._preview_worker = PreviewWorker(self._importer.import_mat)
        self._preview_worker.moveToThread(self._preview_thread)
        self._preview_thread.finished.connect(self._preview_worker.deleteLater)
        self.get_preview.connect(self._preview_worker.create_preview_table)
        self._preview_worker.preview_ready.connect(self.set_preview_table)
        self._preview_thread.start()
        self.collect_preview_values()

    @qt_compat.Slot()
    def collect_preview_values(self):
        """Collect preview data from mat file."""
        no_rows = self.no_preview_rows.value
        row_offset = self.data_offset.value - 1
        if self._valid:
            self.get_preview.emit(no_rows, 0, row_offset, -1, -1, -1)
        else:
            self.data_table = table.File()
        self.update_table.emit()

    @qt_compat.Slot(table.File)
    def set_preview_table(self, data_table):
        self.data_table = data_table
        self.update_table.emit()

    def cleanup(self):
        self._preview_thread.quit()
        self._preview_thread.wait()


class ImporterXLS(object):
    """Class for importat of data in the xlsx format."""

    def __init__(self, fq_infilename, multi):
        self._warnings = set()
        self._multi = multi
        self._fq_infilename = fq_infilename

        def raise_mixed_types(x):
            class MixedTypesError(Exception):
                pass

            raise MixedTypesError()

        def type_converters(allowed_types):
            """
            Takes a dict with convertable types and returns a new dict with all
            types represented. Unallowed types will be filled in with
            raise_mixed_types, except for NoneType which is always allowed and
            results in a None value.
            """
            defaults = {
                str: raise_mixed_types,
                bool: raise_mixed_types,
                int: raise_mixed_types,
                float: raise_mixed_types,
                datetime.datetime: raise_mixed_types,
                datetime.time: raise_mixed_types,
                type(None): lambda x: None,
            }
            return {**defaults, **allowed_types}

        def float_to_float(x):
            # Excel numbers never have more than 15 significant digits
            # so anything beyond that is just an effect of conversion
            # from excel format to python floats.
            return round(x, 15)

        def date_to_date(x):
            if x < datetime.datetime(1900, 3, 1):
                self._warnings.add(
                    "Ignoring any datetimes before 1900-03-01. All dates before "
                    "1900-03-01 are ambiguous because of a bug in Excel.")
                return None
            return x

        def date_to_str(x):
            date = date_to_date(x)
            if date is None:
                return None
            else:
                return str(date)

        def time_to_timedelta(x):
            return datetime.timedelta(
                hours=x.hour,
                minutes=x.minute,
                seconds=x.second,
                microseconds=x.microsecond)

        self._xl_cell_to_str = type_converters({
            str: str,
            bool: lambda x: str(bool(x)),
            int: str,
            float: lambda x: str(float_to_float(x)),
            datetime.datetime: date_to_str,
            datetime.time: lambda x: str(time_to_timedelta(x)),
            type(None): lambda x: '',
        })
        self._xl_cell_to_bool = type_converters({
            bool: bool,
        })
        self._xl_cell_to_int = type_converters({
            bool: int,
            int: int,
        })
        self._xl_cell_to_float = type_converters({
            bool: float,
            int: float,
            float: float_to_float,
        })
        self._xl_cell_to_date = type_converters({
            datetime.datetime: date_to_date,
        })
        self._xl_cell_to_time = type_converters({
            datetime.time: time_to_timedelta,
        })

    def get_row_to_array(
            self, ws, row, data_start_column=0, data_end_column=None):
        """Read a row in the xls file and returns it as an array."""
        return self._to_array(ws, 'row', row,
                              data_start=data_start_column,
                              data_end=data_end_column)

    def get_column_to_array(
            self, ws, column, data_start_row=0, data_end_row=None):
        """Read a column in the xls file and returns it as an array."""
        return self._to_array(ws, 'column', column,
                              data_start=data_start_row,
                              data_end=data_end_row)

    def _to_array(self, ws, alignment, i, data_start=0, data_end=None):
        """Read a row/column in the xls file and returns it as an array."""
        i += 1
        data_start += 1
        if data_end is not None:
            data_end += 1

        if alignment == 'row':
            iter_values = ws.iter_rows
            min_row=i
            max_row=i
            min_col=data_start
            max_col=data_end or ws.max_column
        elif alignment == 'column':
            iter_values = ws.iter_cols
            min_col=i
            max_col=i
            min_row=data_start
            max_row=data_end or ws.max_row
        else:
            raise ValueError("Invalid alignment: {}. Should be either "
                             "'row' or 'column'.".format(alignment))
        min_row = max(min_row, 1)
        max_row = min(max_row, ws.max_row)
        min_col = max(min_col, 1)
        max_col = min(max_col, ws.max_column)
        values_gen = iter_values(
            min_col=min_col,
            max_col=max_col,
            min_row=min_row,
            max_row=max_row,
            values_only=True,
        )
        if values_gen == tuple():
            # openpyxl documentation states that iter_rows/iter_cols can return
            # an empty tuple instead of a generator if there are no cells in
            # the worksheet. This doesn't seem to actually be the case in
            # openpyxl==3.0.2, but it does happen for Worksheet.columns so this
            # is just to be on the safe side.
            values = []
        else:
            values = list(next(values_gen))

        types = {type(v) for v in values}  - {type(None)}
        if len(types) == 0:
            # Only emtpy cells in this row/column => convert to str
            target_type = str
        elif len(types) == 1:
            # Only a single type => convert to that type
            target_type = list(types)[0]
        else:
            if not types - {int, bool}:
                # Only integer types (including bool) => convert all to int
                target_type = int
            elif not types - {float, int, bool}:
                # Only numeric types (including bool) => convert all to float
                target_type = float
            else:
                # A combination of str/datetime/numeric => convert all to str
                target_type = str

        # Convert all values to target type
        converter = {
            str: self._xl_cell_to_str,
            int: self._xl_cell_to_int,
            float: self._xl_cell_to_float,
            datetime.datetime: self._xl_cell_to_date,
            datetime.time: self._xl_cell_to_time,
            bool: self._xl_cell_to_bool,
        }[target_type]
        array_values = [converter[type(v)](v) for v in values]
        mask = [v is None for v in array_values]

        if any(mask):
            # Deal with empty cells in row/column
            if target_type == datetime.datetime:
                empty_value = datetime.datetime(1900, 1, 1)
            else:
                empty_value = target_type()
            array_values = [empty_value if v is None else v
                            for v in array_values]
            return np.ma.MaskedArray(array_values, mask=mask)
        else:
            return np.array(array_values)

    def import_xls(self, out_data, sheet_name, import_all,
                   data_start, data_end,
                   transposed, headers_row_offset=-1,
                   units_row_offset=-1,
                   descriptions_row_offset=-1):
        """Adminstration method for import of data in xls format."""

        def sheet_by_name(sheet_name):
            try:
                ws = wb[sheet_name]
            except KeyError:
                # This is used if called from the auto importer or if the
                # config is incorrect.
                ws = wb.worksheets[0]
            return ws

        def import_sheet(ws, out_table):
            out_table.set_name(ws.title)

            if transposed:
                nr_rows = ws.max_column
                nr_cols = ws.max_row
                data_collector = self.get_row_to_array
                header_collector = self.get_column_to_array
            else:
                nr_rows = ws.max_row
                nr_cols = ws.max_column
                data_collector = self.get_column_to_array
                header_collector = self.get_row_to_array

            if data_end < 0:
                data_end_ = nr_rows + data_end
            elif data_end == 0:
                data_end_ = nr_rows
            else:
                data_end_ = data_start + data_end

            if headers_row_offset >= 0:
                headers = header_collector(ws, headers_row_offset)
                if isinstance(headers, np.ma.MaskedArray):
                    headers = headers.filled('')
            else:
                headers = ['f{0}'.format(index) for index in range(nr_cols)]

            if units_row_offset >= 0:
                units = [str(x) for x in
                         header_collector(ws, units_row_offset)]
            else:
                units = None

            if descriptions_row_offset >= 0:
                descs = [str(x) for x in
                         header_collector(ws, descriptions_row_offset)]
            else:
                descs = None

            for col in range(nr_cols):
                if not headers[col]:
                    continue
                attr = {}
                out_table.set_column_from_array(
                    headers[col], data_collector(
                        ws, col, data_start, data_end_))
                if units:
                    attr['unit'] = units[col]
                if descs:
                    attr['description'] = descs[col]

                out_table.set_column_attributes(headers[col], attr)

        if not os.path.exists(self._fq_infilename):
            raise SyDataError(
                "File does not exist: {}".format(self._fq_infilename))
        elif is_xls(self._fq_infilename):
            raise SyDataError(
                "Support for legacy Excel file format (XLS) has been removed "
                "from Sympathy 2.0.0. Please use Excel or another external "
                "tool to convert the file to the newer Excel file format "
                "(XLSX).")

        with open_workbook(self._fq_infilename) as wb:
            if self._multi:
                if import_all:
                    for ws in wb.worksheets:
                        out_table = out_data.create()
                        import_sheet(ws, out_table)
                        out_data.append(out_table)
                else:
                    out_table = out_data.create()
                    import_sheet(sheet_by_name(sheet_name), out_table)
                    out_data.append(out_table)
            else:
                out_table = out_data
                import_sheet(sheet_by_name(sheet_name), out_table)

        for warning in self._warnings:
            sywarn(warning)


class ImporterCSV(object):
    """Importer class for data in csv format."""

    def __init__(self, csv_source):
        self._source = csv_source
        self._partial_read = False

    def set_partial_read(self, value):
        self._partial_read = value

    def import_csv(self, out_table,
                   nr_data_rows, data_foot_rows, data_row_offset,
                   headers_row_offset=-1,
                   units_row_offset=-1,
                   descriptions_row_offset=-1,
                   require_num=True,
                   read_csv_full_rows=False):
        """Adminstration method for import of data in csv format."""
        self._discard = False
        self._require_num = require_num

        data_table = self._data_as_table(
            nr_data_rows, data_foot_rows, data_row_offset,
            read_csv_full_rows)

        if read_csv_full_rows:
            out_table.update(data_table)
            return

        if headers_row_offset >= 0:
            header_table = self._collect_single_row(headers_row_offset)
            unit_table = table.File()
            desc_table = table.File()

            if units_row_offset >= 0:
                unit_table = self._collect_single_row(units_row_offset)
            if descriptions_row_offset >= 0:
                desc_table = self._collect_single_row(descriptions_row_offset)

            for column_name in header_table.column_names():
                header = str(
                    header_table.get_column_to_array(column_name)[0])
                attr = {}

                if header:

                    try:
                        out_table.update_column(
                            header, data_table, column_name)
                    except KeyError:
                        empty_column = np.array(
                            [u''] * data_table.number_of_rows())
                        out_table.set_column_from_array(header, empty_column)

                    try:
                        unit = unit_table.get_column_to_array(column_name)[0]
                        attr['unit'] = str(unit)
                    except KeyError:
                        pass

                    try:
                        description = desc_table.get_column_to_array(
                            column_name)[0]
                        attr['description'] = str(description)
                    except KeyError:
                        pass

                    if attr:
                        out_table.set_column_attributes(
                            header, attr)

                else:
                    pass

        else:
            out_table.update(data_table)

    def _collect_single_row(self, row_offset):
        """
        Import a single line from the csv-file. This method is used for
        header, units and descriptions.
        """
        no_rows = 1
        require_num = False
        return self._source.read_part(no_rows, row_offset, require_num)

    def _data_as_table(self, no_rows, foot_rows, offset_rows, full_rows):
        """
        Merges the imported data, stored in one or many Tables, into a single
        Table.
        """
        out_table = table.File()

        if full_rows:
            table_list = self._data_as_tables_full_rows(
                no_rows, offset_rows)
        else:
            table_list = self._data_as_tables(
                no_rows, foot_rows, offset_rows)

        if len(table_list) > 1:
            out_table.vjoin(table_list, '', '', True, 0)
        elif len(table_list) == 1:
            out_table = table_list[0]

        if foot_rows > 0 and not self._discard:
            return out_table[:-foot_rows]
        else:
            return out_table

    def _data_as_tables(self, no_rows, foot_rows, offset_rows, iter_count=0):
        """
        Import data from csv-file in chunks. Each chunk is represented
        by a Table in the list, table_list.
        """
        table_list = []
        iter_count += 1

        if iter_count > ITER_LIMIT:
            message = ('Process has ended because the number of calls '
                       'of the method "_data_as_tables" have passed the '
                       'allowed limit, {0}'.format(ITER_LIMIT))
            raise Exception(message)

        if no_rows <= 0:
            try:
                table_list.append(self._source.read_to_end_no_chunks(
                    offset_rows, self._require_num))
            except TooManyColumnsError as tmce:
                table_list.extend(self._split_reading_whole_file(
                    tmce.line, offset_rows, iter_count))

        elif no_rows > 0:
            try:
                table_list.append(self._source.read_part(
                    no_rows, offset_rows, self._require_num))
            except TooManyColumnsError as tmce:
                table_list.extend(self._split_reading_part_file(
                    tmce.line, offset_rows, no_rows, iter_count))
        else:
            raise Exception('Not valid number of rows to read.')

        return table_list

    def _data_as_tables_full_rows(self, no_rows, offset_rows):
        """Import data from csv-file as whole rows."""
        if no_rows == -1:
            return self._source.read_to_end_full_rows(offset_rows)
        elif no_rows > 0:
            return self._source.read_part_full_rows(no_rows, offset_rows)
        else:
            raise Exception('Not valid number of rows to read.')

    def _split_reading_whole_file(self, line, offset_rows, iter_count):
        """Method called if the number of columns in a row is higher than
        expected by pandas.io.parser. The reading from the csv file is
        splited into two readings. This method is used when the whole
        csv file is imported.
        """
        out_list = []

        no_rows_1 = line - 1 - offset_rows
        offset_1 = offset_rows
        offset_2 = line - 1

        if offset_1 == offset_2:
            self._discard_warning(offset_1)
            return out_list

        out_list.extend(
            self._data_as_tables(no_rows_1, 0, offset_1, iter_count))

        out_list.extend(
            self._data_as_tables(-1, 0, offset_2, iter_count))

        return out_list

    def _split_reading_part_file(self, line, offset_rows,
                                 no_rows, iter_count):
        """Method called if the number of columns in a row is higher than
        expected by pandas.io.parser. The reading from the csv file is
        splited into two readings. This method is used when only a part of
        the csv file is imported.
        """
        out_list = []

        no_rows_1 = line - 1 - offset_rows
        offset_1 = offset_rows
        no_rows_2 = no_rows - no_rows_1
        offset_2 = line - 1

        if offset_1 == offset_2:
            self._discard_warning(offset_1)
            return out_list
        if (no_rows + offset_rows) < line:
            raise SyDataError(
                'File is corrupt, error in row: {}'.format(line))

        out_list.extend(
            self._data_as_tables(no_rows_1, 0, offset_1, iter_count))

        out_list.extend(
            self._data_as_tables(no_rows_2, 0, offset_2, iter_count))

        return out_list

    def _discard_warning(self, row):
        if self._partial_read:
            sywarn(
                'Error in row: {}, discarding it and every row below.'.format(
                    row + 1))
            self._discard = True
        else:
            raise SyDataError(
                'File is corrupt, error in row: {}'.format(row + 1))


class ImporterMAT(object):
    """Importer class for data in mat format."""

    def __init__(self, mat_source):
        self._source = mat_source
        self._partial_read = False

    def set_partial_read(self, value):
        self._partial_read = value

    def import_mat(self, out_table, nr_data_rows, data_foot_rows,
                   data_row_offset, units_row_offset=-1,
                   require_num=True,
                   read_mat_full_rows=False):

        self._discard = False

        data_table = self._data_as_table(nr_data_rows, data_foot_rows,
                                         data_row_offset, read_mat_full_rows)

        out_table.update(data_table)

    def _data_as_table(self, no_rows, foot_rows, offset_rows, full_rows):
        """
        Merges the imported data, stored in one or many Tables, into a single
        Table.
        """
        out_table = table.File()

        if full_rows:
            table_list = self._data_as_tables_full_rows(no_rows, offset_rows)
        else:
            table_list = self._data_as_tables(no_rows, foot_rows, offset_rows)

        if len(table_list) > 1:
            out_table.vjoin(table_list, '', '', True, 0)
        elif len(table_list) == 1:
            out_table = table_list[0]

        if foot_rows > 0 and not self._discard:
            return out_table[:-foot_rows]
        else:
            return out_table

    def _data_as_tables(self, no_rows, foot_rows, offset_rows, iter_count=0):
        """
        Import data from mat-file in chunks. Each chunk is represented
        by a Table in the list, table_list.
        """
        table_list = []
        iter_count += 1

        if iter_count > ITER_LIMIT:
            message = ('Process has ended because the number of calls '
                       'of the method "_data_as_tables" have passed the '
                       'allowed limit, {0}'.format(ITER_LIMIT))
            raise Exception(message)

        if no_rows < 0:
            table_list.append(self._source.read(0, offset_rows, False))
        elif no_rows >= 0:
            table_list.append(self._source.read(no_rows, offset_rows, True))
        else:
            raise Exception('Not valid number of rows to read.')

        return table_list

    def _data_as_tables_full_rows(self, no_rows, offset_rows):
        """Import data from mat-file as whole rows."""
        if no_rows == -1:
            return self._source.read(-1, offset_rows, True)
        elif no_rows > 0:
            return [(self._source.read(no_rows, offset_rows, True))]
        else:
            raise Exception('Not valid number of rows to read.')


class TableSourceCSV(object):
    DEFAULT_DELIMITER = ';'
    DEFAULT_ENCODING = 'ascii'
    """
    This class is the layer between the physical csv-file and the import
    routines.
    """

    def __init__(self, fq_infilename, delimiter=None, encoding=None,
                 valid=True):
        self._fq_infilename = fq_infilename
        self._delimiter = self.DEFAULT_DELIMITER
        self._encoding = self.DEFAULT_ENCODING
        self._chunksize = None
        self._valid_encoding = None

        if valid:
            if encoding is None:
                self._encoding = self._sniff_encoding()
            else:
                self._valid_encoding = self._check_encoding(
                    encoding, delimiter)
                self._encoding = encoding

            if delimiter is None:
                self._delimiter = self._sniff_delimiter()
            else:
                self._delimiter = str(delimiter)

        self._has_header = self._sniff_header()

        self._no_rows = None

    def filesize(self):
        """Return filesize of csv-file."""
        return os.path.getsize(self._fq_infilename)

    @property
    def nrows(self):
        """Return the number of row in csv-file, call counter
        if value is not known.
        """
        if self._no_rows is None:
            self._no_rows = self._row_counter()
        return self._no_rows

    @property
    def delimiter(self):
        """Return present delimiter."""
        return self._delimiter

    @delimiter.setter
    def delimiter(self, delimiter):
        """Set delimiter fromthe outside."""
        self._delimiter = str(delimiter)

    @property
    def encoding(self):
        """Return present encoding."""
        return self._encoding

    @encoding.setter
    def encoding(self, encoding):
        """Set encoding of file from outside the source class."""
        self._encoding = encoding
        self._valid_encoding = self._check_encoding(
            self._encoding, self._delimiter)
        self._no_rows = None

    @property
    def chunksize(self):
        """Return present encoding."""
        return self._chunksize

    @chunksize.setter
    def chunksize(self, chunksize):
        """Set chunksize when reading from outside the source class."""
        self._chunksize = chunksize

    @property
    def object(self):
        """Open and return csv-file as file object."""
        self._csvfile = codecs.open(
            self._fq_infilename, mode='rb', encoding=self.encoding)
        return self._csvfile

    def close_object(self):
        """Close open csv-file."""
        self._csvfile.close()

    @property
    def has_header(self):
        """Return True if if header has been found in csv-file."""
        return self._has_header

    def read_to_end_full_rows(self, offset):
        """Quick call for import of data in csv-file as full rows."""
        return self._read_full_rows(-1, offset)

    def read_part_full_rows(self, no_rows, offset):
        """
        Quick call for import of a part of data in csv-file
        as full rows.
        """
        return self._read_full_rows(no_rows, offset)

    def read_to_end_no_chunks(self, offset, req_num):
        """
        Quick call for read the whole csv-file. The method gets
        a dataframe from _read and return a table to caller.
        """
        return table.File().from_dataframe(
            self._read(row_offset=offset,
                       require_num=req_num).dropna(how='all'))

    def read_part(self, no_rows, offset, req_num):
        """
        Quick call for just read a part of the csv-file. The method gets
        a dataframe from _read and return a table to caller.
        """
        return table.File().from_dataframe(
            self._read(no_rows=no_rows,
                       row_offset=offset,
                       require_num=req_num).dropna(how='all'))

    def _read(self, no_rows=None, row_offset=None, foot_rows=None,
              require_num=True, chunksize=None):
        """Administration of the reading of csv-files with pandas."""
        encoding = self.encoding
        delimiter = self._delimiter

        # Detect utf_8 BOM and change encoding if it is there.
        if encoding == 'utf_8':
            with open(self._fq_infilename, 'rb') as f:
                if f.read(len(codecs.BOM_UTF8)) == codecs.BOM_UTF8:
                    encoding = 'utf_8_sig'

        if delimiter != '':
            if encoding.startswith('utf_16'):
                return self._read_python(
                    no_rows, row_offset, encoding,
                    delimiter, foot_rows, chunksize)
            else:
                return self._read_c(
                    no_rows, row_offset, encoding,
                    delimiter, foot_rows, require_num, chunksize)
        else:
            if chunksize is not None:
                return [pd.DataFrame()]
            else:
                return pd.DataFrame()

    def _read_sniff(self, no_rows=None, row_offset=None,
                    encoding=None, delimiter=None):
        """Quick call used by the sniffer methods."""
        chunksize = None
        foot_rows = 0

        return self._read_python(
            no_rows, row_offset, encoding, delimiter, foot_rows, chunksize)

    def _read_c(self, no_rows, row_offset, encoding,
                delimiter, foot_rows, require_num, chunksize):
        """Method for reading csv file with pandas.io.parsers.read_csv,
        c engine.
        """
        foot_rows = foot_rows or 0

        try:
            if StrictVersion(pd.__version__) >= StrictVersion('0.15.0'):
                return PANDAS_CSV_READER(self._fq_infilename,
                                         sep=delimiter,
                                         skipinitialspace=True,
                                         header=None,
                                         prefix='X',
                                         skiprows=row_offset,
                                         skipfooter=foot_rows,
                                         nrows=no_rows,
                                         chunksize=chunksize,
                                         encoding=encoding,
                                         engine='c',
                                         na_filter=require_num,
                                         skip_blank_lines=False)
            else:
                sywarn(
                    'The csv import routine is optimized to work '
                    'with pandas v.0.15.0 or newer. '
                    'Your present verison is {}'.format(pd.__version__))
                return PANDAS_CSV_READER(self._fq_infilename,
                                         sep=delimiter,
                                         skipinitialspace=True,
                                         header=None,
                                         prefix='X',
                                         skiprows=row_offset,
                                         skipfooter=foot_rows,
                                         nrows=no_rows,
                                         chunksize=chunksize,
                                         encoding=encoding,
                                         engine='c',
                                         na_filter=require_num)

        except Exception as message:
            errorline = re.match(TOO_MANY_COLUMNS, str(message))

            if errorline:
                raise TooManyColumnsError(
                    message, int(errorline.groups(0)[0]))
            else:
                errorline = re.match(END_OF_LINE, str(message))
                if errorline:
                    raise TooManyColumnsError(
                        message, int(errorline.groups(0)[0]) + 1)
                else:
                    return self._read_python(no_rows,
                                             row_offset,
                                             encoding,
                                             delimiter,
                                             foot_rows,
                                             chunksize)

    def _read_python(
            self, no_rows, row_offset, encoding,
            delimiter, foot_rows, chunksize):
        """Method for reading csv file with pandas.io.parsers.read_csv,
        python engine.
        """
        foot_rows = foot_rows or 0

        try:
            if StrictVersion(pd.__version__) >= StrictVersion('0.15.0'):
                return PANDAS_CSV_READER(self._fq_infilename,
                                         sep=delimiter,
                                         skipinitialspace=True,
                                         header=None,
                                         prefix='X',
                                         skiprows=row_offset,
                                         skipfooter=foot_rows,
                                         nrows=no_rows,
                                         chunksize=chunksize,
                                         encoding=encoding,
                                         engine='python',
                                         skip_blank_lines=False)
            else:
                sywarn(
                    'The csv import routine is optimized to work '
                    'with pandas v.0.15.0 or newer. '
                    'Your present verison is {}'.format(pd.__version__))
                return PANDAS_CSV_READER(self._fq_infilename,
                                         sep=delimiter,
                                         skipinitialspace=True,
                                         header=None,
                                         prefix='X',
                                         skiprows=row_offset,
                                         skipfooter=foot_rows,
                                         nrows=no_rows,
                                         chunksize=chunksize,
                                         encoding=encoding,
                                         engine='python')

        except Exception as message:
            errorline = re.match(TOO_MANY_COLUMNS, str(message))

            if errorline:
                raise TooManyColumnsError(
                    message, int(errorline.groups(0)[0]))
            else:
                errorline = re.match(END_OF_LINE, str(message))
                if errorline:
                    raise TooManyColumnsError(
                        message, int(errorline.groups(0)[0]) + 1)
                else:
                    raise

    def _read_full_rows(self, no_rows, offset):
        encoding = self.encoding
        table_list = []

        with codecs.open(self._fq_infilename,
                         mode='rb',
                         encoding=encoding) as csvfile:

            try:
                for row in range(offset):
                    next(csvfile)
            except StopIteration:
                return table_list

            if no_rows > 0:
                out_table = table.File()
                data_list = []
                try:
                    for x in range(no_rows):
                        data_list.append(next(csvfile).strip())
                except StopIteration:
                    return table_list
                out_table.set_column_from_array(
                    'X0', np.array(data_list))
                table_list.append(out_table)
            else:
                byte_to_end = self.filesize() - csvfile.tell()
                while byte_to_end > CHUNK_BYTE_LIMIT:
                    out_table = table.File()
                    data_list = []
                    for x in range(CHUNK_ROW_LIMIT):
                        data_list.append(next(csvfile).strip())

                    out_table.set_column_from_array(
                        'X0', np.array(data_list))
                    table_list.append(out_table)
                    byte_to_end = self.filesize() - csvfile.tell()

                out_table = table.File()
                data = np.array(csvfile.read().split('\n'))
                if data[-1] == '':
                    data = data[:-1]

                if len(data):
                    out_table.set_column_from_array('X0', data)
                table_list.append(out_table)

        return table_list

    @property
    def valid_file(self):
        """Return True if file is a csv-file."""
        return self._check_file_validity()

    @property
    def valid_encoding(self):
        """Return True if valid encoding has been determined."""
        return self._valid_encoding

    def _check_file_validity(self):
        """Check if incoming file is a valid csv-file."""
        not_allowed_extensions = ['xls', 'xlsx']
        extension = os.path.splitext(self._fq_infilename)[1][1:]
        if extension in not_allowed_extensions:
            return False
        else:
            return self._valid_encoding

    def _row_counter(self):
        """Count the number of rows in csv-file."""
        def blocks(files, size=65536):
            while True:
                b = files.read(size)
                if not b:
                    break
                yield b

        if self._valid_encoding:
            line_nr = 0
            with open(self._fq_infilename, 'rb') as f:
                line_nr = sum(bl.count('\n') for bl in blocks(f))
            return line_nr + 1
        else:
            return -1

    def _sniff_encoding(self):
        """Sniff if either utf_8, iso8859_1 or utf_16 is a valid encodings."""
        default_encoding = 'utf_8'
        # encodings = ['utf_8', 'utf_16', 'iso8859_1']
        encodings = ['utf_8', 'utf_16', 'utf_16_le', 'utf_16_be', 'iso8859_1']
        # First, test without delimiter.
        for encoding in encodings:
            if self._check_encoding(encoding):
                self._valid_encoding = True
                return encoding

        self._valid_encoding = False
        return default_encoding

    def _check_encoding(self, encoding, delimiter=None):
        """Check if incoming encoding is valid."""
        def check_small_file(csvfile, check=True):
            return checker(csvfile.read(), check)

        def check_large_file(csvfile, encoding, check=True):
            valid_start = checker(csvfile.read(SNIFF_LIMIT))
            if encoding != 'utf_16':
                csvfile.seek(-SNIFF_LIMIT, os.SEEK_END)
                valid_end = checker(csvfile.read(), check)
            else:
                valid_end = True

            return (valid_start and valid_end)

        def checker(sniff_text, check=True):
            if not check:
                return True

            row_split = sniff_text.split('\n')
            if len(row_split) > 1:
                return True
            return False

        def check_file():
            return valid_encoding

        valid_encoding = True
        has_newlines = False

        with open(self._fq_infilename, 'rb') as csvfile:
            data = csvfile.readline()
            if data == b'':
                return True

            has_newlines = b'\n' in data

        with codecs.open(self._fq_infilename,
                         mode='rb',
                         encoding=encoding) as csvfile:
            try:
                if self.filesize() > SNIFF_LIMIT:
                    valid_encoding = check_large_file(
                        csvfile, encoding, has_newlines)
                else:
                    valid_encoding = check_small_file(csvfile, has_newlines)
            except Exception:
                valid_encoding = False

        return valid_encoding

    def _sniff_delimiter(self):
        """Method tries to determine a valid delimiter."""
        if self._valid_encoding:
            best = self.DEFAULT_DELIMITER
            best_count = 0
            smallest_diff = 1000
            num_lines = 50
            encoding = self._encoding

            delimiters = [
                value for key, value in CSV_FILE_DELIMITERS.items()
                if key != 'Other']
            counts = []

            for delimiter in delimiters:
                try:
                    # First row not included. High probability to
                    # be the header row which can include a lot of spaces
                    counts.append(self._read_sniff(
                        no_rows=num_lines,
                        row_offset=1,
                        encoding=encoding,
                        delimiter=delimiter).shape[1])
                except Exception:
                    counts.append(-1)
            try:
                count, delimiter = max(zip(counts, delimiters))
                if count and count > 1:
                    return delimiter
            except ValueError:
                pass

            # Slower fallback method of determining delimiter if the
            # fast one fails to give a delimiter.

            for delimiter in delimiters:
                counts = []

                for offset in range(num_lines):
                    try:
                        data_frame = self._read_sniff(
                            no_rows=1,
                            row_offset=offset,
                            encoding=encoding,
                            delimiter=delimiter)
                        counts.append(data_frame.shape[1])
                    except Exception:
                        pass

                if not counts:
                    continue

                if (abs(min(counts) - max(counts)) < smallest_diff and
                        max(counts) > 1):
                    best = delimiter
                    best_count = max(counts)
                    smallest_diff = abs(min(counts) - max(counts))
                elif (abs(min(counts) - max(counts)) == smallest_diff and
                        max(counts) > best_count):
                    best = delimiter
                    best_count = max(counts)
                    smallest_diff = abs(min(counts) - max(counts))
                else:
                    pass

            return best
        else:
            return str(self.DEFAULT_DELIMITER)

    def _sniff_header(self):
        """Method tries to determine if there exist a header in csv-file."""
        if self._valid_encoding:
            try:
                header_table = self.read_part(1, 0, False)
                data_table = self.read_part(20, 1, True)
            except Exception:
                return False

            if data_table.number_of_rows() < 1:
                return False

            if (header_table.number_of_columns() <
                    data_table.number_of_columns()):
                return False

            has_header = 0
            for column in header_table.column_names():
                header_array = header_table.get_column_to_array(column)
                try:
                    data_array = data_table.get_column_to_array(column)
                except KeyError:
                    if header_array[0] == u'':
                        continue
                    else:
                        return False

                if header_array.dtype != data_array.dtype:
                    has_header += 1

            return has_header > 0

        else:
            return False


class TableSourceMAT(object):
    """
    This class is the layer between the physical mat-file and the import
    routines.
    """

    def __init__(self, fq_infilename):
        self._fq_infilename = fq_infilename

    def read(self, no_rows, offset, part):
        out_table = table.File()
        mat_table = matlab.read_matfile_to_table(self._fq_infilename)
        names = mat_table.column_names()
        rows = mat_table.number_of_rows()

        if part:
            end_row = no_rows + offset
            if end_row > rows:
                end_row = rows
        elif no_rows == 0:
            end_row = rows
        else:
            end_row = no_rows

        for column in names:
            data = mat_table.get_column_to_array(column)
            out_table.set_column_from_array(column, data[offset:end_row])

        out_table.set_attributes(mat_table.get_attributes())
        out_table.set_name(mat_table.get_name())
        return out_table


class TooManyColumnsError(Exception):
    """Exception raised when the number of columns are higher than Pandas has
    expected in a row.
    """

    def __init__(self, value, line):
        self._value = value
        self.line = line

    def __str__(self):
        return repr(self._value)
